import pandas as pd
import openpyxl
from openpyxl import load_workbook
import logging
import os
from typing import List, Dict, Any

logger = logging.getLogger(__name__)

class ExcelParser:
    def __init__(self):
        self.supported_extensions = {'.xls', '.xlsx'}
    
    def parse_excel_file(self, file_path: str) -> Dict[str, Any]:
        """
        Parse Excel file and extract text content from all sheets
        Returns a dictionary with extracted text and metadata
        """
        try:
            if not os.path.exists(file_path):
                raise FileNotFoundError(f"File not found: {file_path}")
            
            file_extension = os.path.splitext(file_path)[1].lower()
            if file_extension not in self.supported_extensions:
                raise ValueError(f"Unsupported file type: {file_extension}")
            
            # Extract text content
            extracted_text = self._extract_text_content(file_path)
            
            # Get metadata
            metadata = self._get_file_metadata(file_path)
            
            return {
                'text_content': extracted_text,
                'metadata': metadata,
                'total_sheets': len(metadata['sheets']),
                'total_cells': metadata['total_cells']
            }
            
        except Exception as e:
            logger.error(f"Error parsing Excel file {file_path}: {e}")
            raise Exception(f"Failed to parse Excel file: {str(e)}")
    
    def _extract_text_content(self, file_path: str) -> str:
        """Extract text content from all sheets in the Excel file"""
        workbook = None
        try:
            # Use openpyxl for better text extraction
            workbook = load_workbook(filename=file_path, data_only=True)
            
            extracted_texts = []
            
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                sheet_text = self._extract_sheet_text(sheet, sheet_name)
                if sheet_text.strip():
                    extracted_texts.append(f"=== SHEET: {sheet_name} ===\n{sheet_text}\n")
            
            # Combine all text with intelligent spacing
            combined_text = "\n\n".join(extracted_texts)
            
            # Clean up the text
            cleaned_text = self._clean_text(combined_text)
            
            return cleaned_text
            
        except Exception as e:
            logger.error(f"Error extracting text from {file_path}: {e}")
            raise Exception(f"Text extraction failed: {str(e)}")
        finally:
            # Ensure workbook is closed to release file handle
            if workbook:
                try:
                    workbook.close()
                except:
                    pass
    
    def _extract_sheet_text(self, sheet, sheet_name: str) -> str:
        """Extract text from a single sheet"""
        texts = []
        
        # Get the used range
        max_row = sheet.max_row
        max_col = sheet.max_column
        
        if max_row == 0 or max_col == 0:
            return ""
        
        # Extract text from each cell
        for row in range(1, max_row + 1):
            row_texts = []
            for col in range(1, max_col + 1):
                cell = sheet.cell(row=row, column=col)
                cell_value = cell.value
                
                if cell_value is not None:
                    # Convert to string and clean
                    cell_text = str(cell_value).strip()
                    if cell_text:
                        row_texts.append(cell_text)
            
            # Join row texts with spaces
            if row_texts:
                row_text = " ".join(row_texts)
                texts.append(row_text)
        
        return "\n".join(texts)
    
    def _clean_text(self, text: str) -> str:
        """Clean and normalize extracted text"""
        if not text:
            return ""
        
        # Remove excessive whitespace
        lines = text.split('\n')
        cleaned_lines = []
        
        for line in lines:
            # Remove excessive spaces
            cleaned_line = ' '.join(line.split())
            if cleaned_line:
                cleaned_lines.append(cleaned_line)
        
        # Join lines with proper spacing
        cleaned_text = '\n'.join(cleaned_lines)
        
        # Remove excessive newlines
        while '\n\n\n' in cleaned_text:
            cleaned_text = cleaned_text.replace('\n\n\n', '\n\n')
        
        return cleaned_text.strip()
    
    def _get_file_metadata(self, file_path: str) -> Dict[str, Any]:
        """Get metadata about the Excel file"""
        workbook = None
        try:
            workbook = load_workbook(filename=file_path, data_only=True)
            
            metadata = {
                'filename': os.path.basename(file_path),
                'file_size': os.path.getsize(file_path),
                'sheets': [],
                'total_cells': 0
            }
            
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                max_row = sheet.max_row
                max_col = sheet.max_column
                cell_count = max_row * max_col if max_row > 0 and max_col > 0 else 0
                
                metadata['sheets'].append({
                    'name': sheet_name,
                    'rows': max_row,
                    'columns': max_col,
                    'cells': cell_count
                })
                
                metadata['total_cells'] += cell_count
            
            return metadata
            
        except Exception as e:
            logger.error(f"Error getting metadata for {file_path}: {e}")
            return {
                'filename': os.path.basename(file_path),
                'file_size': os.path.getsize(file_path),
                'sheets': [],
                'total_cells': 0,
                'error': str(e)
            }
        finally:
            # Ensure workbook is closed to release file handle
            if workbook:
                try:
                    workbook.close()
                except:
                    pass
    
    def validate_file(self, file_path: str) -> bool:
        """Validate that the file is a valid Excel file"""
        workbook = None
        try:
            if not os.path.exists(file_path):
                return False
            
            file_extension = os.path.splitext(file_path)[1].lower()
            if file_extension not in self.supported_extensions:
                return False
            
            # Try to open the file to validate it's a proper Excel file
            workbook = load_workbook(filename=file_path, data_only=True)
            return True
            
        except Exception as e:
            logger.error(f"File validation failed for {file_path}: {e}")
            return False
        finally:
            # Ensure workbook is closed to release file handle
            if workbook:
                try:
                    workbook.close()
                except:
                    pass 